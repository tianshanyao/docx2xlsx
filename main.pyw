import openpyxl
import os
import docx
import re
from tkinter import *
import tkinter.filedialog
import tkinter.messagebox
from tkinter import ttk

######### ！！！！网大数据要求：内容是选择性粘贴值的内容，还需人为添加待核查标记########

######### 要求：docx里的文件内容都是选择性粘贴值的内容，保证每一个划分都是段落 #########
######### 要求：docx里题的顺序为：先有题干，再有选项，最后是答案 #######################
######### 结果：重复的题和答案自动去掉，但重复的题、答案不一致的情况会全部保留##########

######### 20180708改进：################################################################
######### 1.改成GUI版本，但仍未封装成类#################################################
######### 2.可以自己选择input文件夹和RE模式#############################################
######### 3.识别.docx文件，其它文件不做处理#############################################

######### 改进想法：是否可以做一个自动产生模式识别的方法？##############################

######### 20180709改进：################################################################
######### 1.添加重复次数################################################################
######### 2.添加单/多选#################################################################
######### 3.捉了点虫：网大版由于格式不一致的问题########################################

######### 20180714改进：################################################################
######### 1.捉了个大虫：question_options只连接了options...##############################



def runmain():
    
    input_path = tkinter.filedialog.askdirectory()
    if not os.path.exists(input_path):
        tkinter.messagebox.showerror('警告', '目录不存在！')
        return
    
    ## 百题通re模式们
    baititong_question_pattern = r'\d{1,3}\.\s'    ## 题干的格式为：1~999.[空格]
    baititong_option_pattern = r'[ABCDEFGHIJKLMN]\.'    ## 选项的格式为：A~N.
    baititong_answer_pattern = r'答案：'    ## 答案的格式为：答案：
    ## L1大学re模式们
    wangda_question_pattern = r'\d{1,3}、\s+'
    wangda_option_pattern = r'\s*[ABCDEFGHIJKLMN]\.\s+'
    wangda_answer_pattern = r'待检查\s*'
    wangda_single_pattern = r'单选'
    wangda_multiple_pattern = r'多选'

    sm_flag = False    ## 是否使用单选/多选re模式，若没有答案，则需要开启单/多选re模式，其它情况则不需要
    
    re_choice = re_text.get()
    ## 选择re模式
    if(re_choice == '百题通'):
        question_pattern = baititong_question_pattern
        option_pattern = baititong_option_pattern
        answer_pattern = baititong_answer_pattern
    elif(re_choice == '网大L考试'):
        question_pattern = wangda_question_pattern
        option_pattern = wangda_option_pattern
        answer_pattern = wangda_answer_pattern
        single_pattern = wangda_single_pattern
        multiple_pattern = wangda_multiple_pattern
        sm_flag = True
    else:
        question_pattern = question_pattern_text.get()
        option_pattern = option_pattern_text.get()
        answer_pattern = answer_pattern_text.get()
    if question_pattern == '' or option_pattern == '' or answer_pattern == '':
        tkinter.messagebox.showerror('警告', 're模式不能为空！')
        return


    
    docx_suffix = '.docx'
    ignore_prefix = '~'
    
    files_ignore = []
    files_read = []
    item_n = 0

    output_xlsx = './result.xlsx'
    output_xls = './result.xls'

    ## 删除原来的result文件们
    if(os.path.exists(output_xlsx)):
        os.remove(output_xlsx)
    if(os.path.exists(output_xls)):
        os.remove(output_xls)

    header = ['序号', '重复次数', '文件来源', '单/多选' , '题干', '答案', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '选项I', '选项J', '选项K', '选项L', '选项M', '选项N']
    max_n_options = 1    ## 选项个数默认为1
    option_header = ['A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.', 'H.', 'G.', 'H.', 'I.', 'J.', 'K.', 'L.', 'M.', 'N.']

    Adict = {}

    question = ''
    answer = ''
    options = []
    question_options = ''
    sm = ''
    no_question = ''

    answers = []

    for (r, ds, fs) in os.walk(input_path):
        for f in fs:
            if f.endswith(docx_suffix) and not f.startswith(ignore_prefix):
                input_docx = '/'.join([input_path, f])
##                print(f)
                docx_file = docx.Document(input_docx)
                for para in docx_file.paragraphs:
                    ## 匹配题干、选项、答案
                    if sm_flag and re.match(single_pattern, para.text):
                        sm = '单选'
                    elif sm_flag and re.match(multiple_pattern, para.text):
                        sm = '多选'
                    elif re.match(question_pattern, para.text):
                        no_question = re.match(question_pattern, para.text).group(0)
                        question = re.sub(question_pattern, '', para.text, 1).strip()
                    elif re.match(option_pattern, para.text):
                        options.append(re.sub(option_pattern, '', para.text, 1).strip())
                    elif re.match(answer_pattern, para.text):
                        answer = re.sub(answer_pattern, '', para.text, 1).strip()
                        if len(answer) == 1:
                            sm = '单选'
                        elif len(answer) > 1:
                            sm = '多选'
                        if(question == '' or not options):
                            question = ''
                            answer = ''
                            options = []
                            question_options = ''
                            no_question = ''
                            continue
                        ## 整理选项和答案，1.将答案从ABC转换成文本答案list，2.将选项排序，3.将答案按排序后的选项换算回ABC
                        ## 答案若先为空会不会有问题？？？？？？？
                        answers = []
                        for a in answer:
                            answers.append(options[ord(a)-ord('A')])
                        options.sort()
                        answer = ''
                        for a in answers:
                            answer = ''.join([answer, chr(ord('A') + options.index(a))])
                        tttt = list(answer)
                        tttt.sort()
                        answer = ''.join(tttt)
                        ## 存到字典中去（选项需加A.格式）
                        ttttt = '-'.join(options)
                        question_options = '-'.join([question, ttttt])
                        while question_options in Adict.keys():    ## 据说in比haskey()方法快
                            if answer != Adict[question_options][5]:
                                question_options = ''.join([question_options, '-'])
                            else:
                                Adict[question_options][1] = Adict[question_options][1] + 1
                                Adict[question_options][2] = ', '.join([Adict[question_options][2], ''.join([no_question, f])])
                                break
                        if question_options not in Adict.keys():
                            Adict[question_options] = [item_n + 1, 1, ''.join([no_question, f]), sm, question, answer, list(map(lambda x, y:''.join([x, y]), option_header[:len(options)], options))]
                            item_n = item_n + 1
##                            print([sm, question, answer])
                        ## 记录max_n_options，重置question、answer、options、question_options
                        if max_n_options < len(options):
                            max_n_options = len(options)
                        question = ''
                        answer = ''
                        options = []
                        question_options = ''
                        no_question = ''
                files_read.append(f)
            else:
                files_ignore.append(f)

    ## 写入xlsx文件
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = input_path.split('/')[-1]
    r = 1
    for c in range(0, max_n_options + 6):
        sheet.cell(row = r, column = c + 1, value = header[c])
    r = 2
    for item in Adict.values():
        for c in range(0, len(item)):
            if type(item[c]) == list:
                for o in range(0, len(item[c])):
                    sheet.cell(row = r, column = c + o + 1, value = item[c][o])
            else:
                sheet.cell(row = r, column = c + 1, value = item[c])
        r = r + 1
    wb.save(output_xlsx)

    result_list = ['保存成功']
    
    result_list.append('共处理文件%d个，处理题目%d个。'%(len(files_read), item_n))
    if files_ignore:
        result_list.append('以下文件未被处理：')
        result_list.extend(files_ignore)
    result_text = '\n\n'.join(result_list)
    tkinter.messagebox.showinfo('提示', result_text)








root = Tk()
root.title('整理考试题 - 从docx到xlsx')

re_text = StringVar()

question_pattern_text = StringVar()
option_pattern_text = StringVar()
answer_pattern_text = StringVar()

attention_text = '''注意：
1. 文件夹里题的顺序为：先有题干，再有选项，最后是答案。题干和选项不可为空，答案可以为空，但必须得有。
2. 文件夹里的文件内容最好是选择性粘贴值的内容，确保每一个划分都是段落。
3. 只识别.docx文件。'''
Label(root, text=attention_text, fg='red').grid(columnspan=3)

Label(root, text='re模式选择：').grid(row=1, sticky=E)
re_choose_combobox = ttk.Combobox(root, textvariable=re_text)
re_choose_combobox['values'] = ('百题通', '网大L考试', '其它')
re_choose_combobox.current(0)
re_choose_combobox.grid(row=1, column=1)
Button(root, text='  执 行  ', command=runmain).grid(row=1, column=2)

Label(root, text='Question_pattern（仅在re选‘其它’时有效）：').grid(row=2, sticky=E)
Entry(root, textvariable=question_pattern_text).grid(row=2, column=1)

Label(root, text='Option_pattern（仅在re选‘其它’时有效）：').grid(row=3, sticky=E)
Entry(root, textvariable=option_pattern_text).grid(row=3, column=1)

Label(root, text='Answer_pattern（仅在re选‘其它’时有效）：').grid(row=4, sticky=E)
Entry(root, textvariable=answer_pattern_text).grid(row=4, column=1)

mainloop()
