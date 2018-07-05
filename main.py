import openpyxl
import os
import docx

######### 要求：docx里的文件内容都是选择性粘贴值的内容，保证每一个划分都是段落 #########
######### 要求：docx里题的顺序为：先有题干，再有选项，最后是答案 #######################
######### 要求：docx里只有docx文件######################################################
######### 结果：重复的题和答案自动去掉，但重复的题、答案不一致的情况会全部保留##########

input_path = '.\\doc'

## 如果input目录不存在，直接退出
if not os.path.exists(input_path):
    print('目录不存在。')
    os.exit()

output_xlsx = './result.xlsx'
##output_xls = './result.xls'

## 删除原来的result文件们
if(os.path.exists(output_xlsx)):
    os.remove(output_xlsx)
if(os.path.exists(output_xls)):
    os.remove(output_xls)

header = ['题干', '答案', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '选项I', '选项J', '选项K', '选项L', '选项M', '选项N']
max_n_options = 1    ## 选项个数默认为1
option_header = ['A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.', 'H.', 'G.', 'H.', 'I.', 'J.', 'K.', 'L.', 'M.', 'N.']

## ------------------------------------------模式们------------------------------------------
## 百题通re模式们
question_pattern = r'\d{1,3}\.\s'    ## 题干的格式为：1~999.[空格]
option_pattern = r'[ABCDEFGHIJKLMN]\.'    ## 选项的格式为：A~N.
answer_pattern = r'答案：'    ## 答案的格式为：答案：
#### L1大学re模式们
##question_pattern = r''
##option_pattern = r''
##answer_pattern = r''





Adict = {}

question = ''
answer = ''
options = []
question_options = ''

answers = []

for (r, ds, fs) in os.walk(input_path):
    for f in fs:
        input_docx = '\\'.join([input_path, f])
        docx_file = docx.Document(input_docx)
        for para in docx_file.paragraphs:
            ## 匹配题干、选项、答案
            if re.match(question_pattern, para.text):
                question = re.sub(question_pattern, '', para.text, 1)
            elif re.match(option_pattern, para.text):
                options.append(re.sub(option_pattern, '', para.text, 1))
            elif re.match(answer_pattern, para.text):
                answer = re.sub(answer_pattern, '', para.text, 1)
                if(question == '' or not options):
                    question = ''
                    answer = ''
                    options = []
                    question_options = ''
                    continue
                ## 整理选项和答案，1.将答案从ABC转换成文本答案list，2.将选项排序，3.将答案按排序后的选项换算回ABC
                ## 答案若先为空会不会有问题？？？？？？？
                answers = []
                for a in answer:
                    answers.append(options[ord(a)-ord('A')])
                options.sort()
                answer = ''
                for a in answers:
                    answer = ''.join([answer, chr(ord('A') + options.index(a))])
                ## 存到字典中去（选项需加A.格式）
                question_options = '-'.join(question.extend(options))
                while question_options in Adict.keys():    ## 据说in比haskey()方法快
                    if answer != Adict[question_options][1]:
                        question_options = ''.join([question_options, '-'])
                    else:
                        break
                if question_options not in Adict.keys():
                    Adict[question_options] = [question, answer, list(map(lambda x, y:''.join([x, y]), option_header[:len(options)], options))]
                ## 记录max_n_options，重置question、answer、options、question_options
                if max_n_options < len(options):
                    max_n_options = len(options)
                question = ''
                answer = ''
                options = []
                question_options = ''









#########################################################
##Adict = {'question-Aoption-Boption-Coption':['question','answer','Aoption','Boption','Coption'],
##         'question2-Aoption2-Boption2-Coption2-Doption2':['question2','answer2','Aoption2','Boption2','Coption2','Doption2']}
##
##max_n_options = 4
##
##
##
##
##
########################################################################







## 写入xlsx文件
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = input_path[2:]
r = 1
for c in range(0, max_n_options + 2):
    sheet.cell(row = r, column = c + 1, value = header[c])
r = 2
for item in Adict.values():
    for c in range(0, len(item)):
        sheet.cell(row = r, column = c + 1, value = item[c])
    r = r + 1
wb.save(output_xlsx)
print('保存成功')
